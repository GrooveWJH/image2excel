from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

def column_letter(idx):
    """Convert a column index to an Excel column letter."""
    letters = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

def resize_image(image, max_width, max_height):
    """Resize image to fit within max_width and max_height while maintaining aspect ratio."""
    width, height = image.size
    aspect_ratio = width / height

    if width > max_width or height > max_height:
        if width / max_width > height / max_height:
            width = max_width
            height = int(width / aspect_ratio)
        else:
            height = max_height
            width = int(height * aspect_ratio)

    return image.resize((width, height), Image.LANCZOS)

def image_to_excel(image_path, output_excel, max_width=256, max_height=256):
    # 打开图像并调整其大小
    img = Image.open(image_path)
    img = resize_image(img, max_width, max_height)
    width, height = img.size
    
    # 创建一个新的工作簿和活动工作表
    wb = Workbook()
    ws = wb.active

    # 设置每个单元格的大小
    for i in range(1, width + 1):
        col_letter = column_letter(i)
        ws.column_dimensions[col_letter].width = 2.7  # 宽度为1单位
    for i in range(1, height + 1):
        ws.row_dimensions[i].height = 15  # 高度为1单位
    
    # 获取每个像素的颜色并填充到相应的单元格
    for y in tqdm(range(height), desc="Processing rows"):
        for x in range(width):
            r, g, b = img.getpixel((x, y))[:3]
            fill = PatternFill(start_color=f'{r:02X}{g:02X}{b:02X}', end_color=f'{r:02X}{g:02X}{b:02X}', fill_type="solid")
            ws.cell(row=y + 1, column=x + 1).fill = fill
    
    print(f"Image converted to Excel successfully!")
    # 保存文件
    wb.save(output_excel)
    print(f"Output Excel file: {output_excel}")
# 示例使用
image_to_excel('input/image.png', 'output/output.xlsx')